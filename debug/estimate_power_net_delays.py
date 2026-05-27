from __future__ import annotations

import csv
import math
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

import openpyxl


SOURCE_CLOCK_MHZ = 175.0
CONFIG_SHEET = "config"
POWER_NET_COUNT = 8
MEASUREMENT_CSV = Path("measured_power_edges.csv")
RESULT_CSV = Path("estimated_power_net_delays.csv")
CURRENT_MODEL_RESULT_CSV = Path("estimated_power_net_delays_current_model.csv")
INFERRED_CDS2_FALL_OFFSET_NS = 0.0


MEASUREMENTS: tuple[tuple[str, str, int], ...] = (
    ("3.3V_DIG", "CDS1↓ to SW↑", -196),
    ("3.3V_DIG", "CDS1↓ to SW↓", 103),
    ("3.3V_DIG", "CDS2↓ to SW↑", 84),
    ("3.3V_DIG", "CDS2↓ to SW↓", -97),
    ("R2.5V_AMP", "CDS1↓ to SW↑", -209),
    ("R2.5V_AMP", "CDS1↓ to SW↓", 151),
    ("R2.5V_AMP", "CDS2↓ to SW↑", 72),
    ("R2.5V_AMP", "CDS2↓ to SW↓", -42),
    ("0.85V_PS", "CDS1↓ to SW↑", -215),
    ("0.85V_PS", "CDS1↓ to SW↓", -137),
    ("0.85V_PS", "CDS2↓ to SW↑", 63),
    ("0.85V_PS", "CDS2↓ to SW↓", 142),
    ("0.85V_PL", "CDS1↓ to SW↑", 35),
    ("0.85V_PL", "CDS1↓ to SW↓", 112),
    ("0.85V_PL", "CDS2↓ to SW↑", -165),
    ("0.85V_PL", "CDS2↓ to SW↓", -88),
    ("1.1V_DDRIO", "CDS1↓ to SW↑", -215),
    ("1.1V_DDRIO", "CDS1↓ to SW↓", -114),
    ("1.1V_DDRIO", "CDS2↓ to SW↑", 65),
    ("1.1V_DDRIO", "CDS2↓ to SW↓", 166),
    ("3.3V_WLAN", "CDS1↓ to SW↑", 89),
    ("3.3V_WLAN", "CDS1↓ to SW↓", -91),
    ("3.3V_WLAN", "CDS2↓ to SW↑", -111),
    ("3.3V_WLAN", "CDS2↓ to SW↓", 189),
    ("4.3V_CHG", "CDS1↓ to SW↑", 111),
    ("4.3V_CHG", "CDS1↓ to SW↓", -194),
    ("4.3V_CHG", "CDS2↓ to SW↑", -89),
    ("4.3V_CHG", "CDS2↓ to SW↓", 87),
    ("1.8V_DIG", "CDS1↓ to SW↑", 114),
    ("1.8V_DIG", "CDS1↓ to SW↓", -202),
    ("1.8V_DIG", "CDS2↓ to SW↑", -87),
    ("1.8V_DIG", "CDS2↓ to SW↓", 77),
    ("5.5V_CHG", "CDS1↓ to SW↑", -110),
    ("5.5V_CHG", "CDS1↓ to SW↓", 238),
    ("5.5V_CHG", "CDS2↓ to SW↑", 171),
    ("5.5V_CHG", "CDS2↓ to SW↓", 39),
    ("D5V_1", "CDS1↓ to SW↑", -150),
    ("D5V_1", "CDS1↓ to SW↓", 237),
    ("D5V_1", "CDS2↓ to SW↑", 130),
    ("D5V_1", "CDS2↓ to SW↓", 37),
    ("D5V_2", "CDS1↓ to SW↑", -145),
    ("D5V_2", "CDS1↓ to SW↓", 239),
    ("D5V_2", "CDS2↓ to SW↑", 135),
    ("D5V_2", "CDS2↓ to SW↓", 39),
    ("D5V_3", "CDS1↓ to SW↑", -139),
    ("D5V_3", "CDS1↓ to SW↓", 236),
    ("D5V_3", "CDS2↓ to SW↑", 142),
    ("D5V_3", "CDS2↓ to SW↓", 36),
    ("A6V", "CDS1↓ to SW↑", -96),
    ("A6V", "CDS1↓ to SW↓", -237),
    ("A6V", "CDS2↓ to SW↑", 182),
    ("A6V", "CDS2↓ to SW↓", 41),
    ("A3V", "CDS1↓ to SW↑", -226),
    ("A3V", "CDS1↓ to SW↓", 145),
    ("A3V", "CDS2↓ to SW↑", 52),
    ("A3V", "CDS2↓ to SW↓", -57),
    ("M15V", "CDS1↓ to SW↑", 163),
    ("M15V", "CDS1↓ to SW↓", -177),
    ("M15V", "CDS2↓ to SW↑", -39),
    ("M15V", "CDS2↓ to SW↓", 103),
    ("A20V", "CDS1↓ to SW↑", 24),
    ("A20V", "CDS1↓ to SW↓", 125),
    ("A20V", "CDS2↓ to SW↑", -179),
    ("A20V", "CDS2↓ to SW↓", -84),
)


@dataclass(frozen=True)
class NetConfig:
    workbook: Path
    index: int
    name: str
    pl_delay_ns: float
    period_ns: float
    duty_percent_min: float
    duty_percent_max: float
    cds1_fall_us: float
    cds2_fall_us: float


def write_measurement_csv() -> None:
    with MEASUREMENT_CSV.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["net", "edge", "measured_ns"])
        writer.writerows(MEASUREMENTS)


def evaluate_formula(value: object, cells: dict[str, object]) -> object:
    if not isinstance(value, str) or not value.startswith("="):
        return value
    formula = value.replace(" ", "").upper()
    match = re.fullmatch(r"=ROUND\(\(1/175\)\*([A-Z]+[1-9][0-9]*)\*10\^3,1\)", formula)
    if match:
        return round((1 / SOURCE_CLOCK_MHZ) * float(cells[match.group(1)]) * 10**3, 1)
    return value


def load_workbook_config(path: Path) -> tuple[dict[str, object], dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[CONFIG_SHEET]
    rows = list(ws.iter_rows())
    headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in rows[0]]
    parameter_col = headers.index("parameter")
    value_col = headers.index("value")
    cells: dict[str, object] = {}
    data: dict[str, object] = {}
    value_cells: dict[str, str] = {}
    for row in rows[1:]:
        key_cell = row[parameter_col]
        value_cell = row[value_col]
        if key_cell.value is None or not str(key_cell.value).strip():
            continue
        key = str(key_cell.value).strip()
        cells[value_cell.coordinate] = value_cell.value
        data[key] = value_cell.value
        value_cells[key] = value_cell.coordinate
    wb.close()
    return {key: evaluate_formula(value, cells) for key, value in data.items()}, value_cells


def load_net_configs() -> dict[str, NetConfig]:
    configs: dict[str, NetConfig] = {}
    for path in sorted(Path(".").glob("シリアル_48us_CLK*.xlsx")):
        data, _ = load_workbook_config(path)
        period_ns = float(data["clock_div_ratio"]) / SOURCE_CLOCK_MHZ * 1000.0
        for index in range(POWER_NET_COUNT):
            name = str(data.get(f"power_net_name_{index}") or "").strip()
            if not name:
                continue
            configs[name] = NetConfig(
                workbook=path,
                index=index,
                name=name,
                pl_delay_ns=float(data["pl_dcdc_clk_delay_ns"]),
                period_ns=period_ns,
                duty_percent_min=float(data[f"power_net_duty_percent_{index}_min"]),
                duty_percent_max=float(data[f"power_net_duty_percent_{index}_max"]),
                cds1_fall_us=float(data["cds1_fall_us"]),
                cds2_fall_us=float(data["cds2_fall_us"]),
            )
    return configs


def nearest_edge_delta_ns(target_us: float, first_edge_ns: float, period_ns: float) -> float:
    target_ns = target_us * 1000.0
    nearest_index = round((target_ns - first_edge_ns) / period_ns)
    candidates = [
        first_edge_ns + index * period_ns
        for index in (nearest_index - 1, nearest_index, nearest_index + 1)
        if index >= 0
    ]
    if not candidates:
        candidates.append(first_edge_ns)
    nearest = min(candidates, key=lambda edge_ns: abs(edge_ns - target_ns))
    return nearest - target_ns


def predicted(config: NetConfig, power_delay_ns: float) -> dict[str, float]:
    rise_ns = config.pl_delay_ns + power_delay_ns
    fall_ns = rise_ns + config.period_ns * config.duty_percent_min / 100.0
    return {
        "CDS1↓ to SW↑": nearest_edge_delta_ns(config.cds1_fall_us, rise_ns, config.period_ns),
        "CDS1↓ to SW↓": nearest_edge_delta_ns(config.cds1_fall_us, fall_ns, config.period_ns),
        "CDS2↓ to SW↑": nearest_edge_delta_ns(config.cds2_fall_us, rise_ns, config.period_ns),
        "CDS2↓ to SW↓": nearest_edge_delta_ns(config.cds2_fall_us, fall_ns, config.period_ns),
    }


def score(config: NetConfig, measured: dict[str, int], power_delay_ns: float) -> float:
    estimates = predicted(config, power_delay_ns)
    return sum((estimates[edge] - value) ** 2 for edge, value in measured.items())


def estimate_delay(config: NetConfig, measured: dict[str, int]) -> tuple[float, dict[str, float]]:
    best_delay = 0.0
    best_score = math.inf
    # The model is periodic, so one clock period contains every distinct solution.
    steps = int(round(config.period_ns * 10))
    for step in range(steps):
        delay = step / 10.0
        value = score(config, measured, delay)
        if value < best_score:
            best_score = value
            best_delay = delay

    # Refine to 0.01 ns around the best 0.1 ns candidate.
    start = max(0.0, best_delay - 0.2)
    end = min(config.period_ns, best_delay + 0.2)
    step_count = int(round((end - start) * 100))
    for step in range(step_count + 1):
        delay = start + step / 100.0
        value = score(config, measured, delay)
        if value < best_score:
            best_score = value
            best_delay = delay
    return round(best_delay, 2), predicted(config, best_delay)


def update_workbooks(results: list[dict[str, object]]) -> None:
    by_workbook: dict[Path, list[dict[str, object]]] = {}
    for row in results:
        by_workbook.setdefault(Path(str(row["workbook"])), []).append(row)

    for path, rows in by_workbook.items():
        backup = path.with_suffix(path.suffix + ".bak")
        if not backup.exists():
            shutil.copy2(path, backup)
        wb = openpyxl.load_workbook(path)
        ws = wb[CONFIG_SHEET]
        _, value_cells = load_workbook_config(path)
        for row in rows:
            key = f"power_net_delay_ns_{row['index']}"
            ws[value_cells[key]] = float(row["estimated_delay_ns"])
        wb.save(path)
        wb.close()


def build_result_rows(
    measured_by_net: dict[str, dict[str, int]],
    configs: dict[str, NetConfig],
    *,
    cds2_fall_offset_ns: float,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for net, measured in measured_by_net.items():
        loaded_config = configs[net]
        config = NetConfig(
            workbook=loaded_config.workbook,
            index=loaded_config.index,
            name=loaded_config.name,
            pl_delay_ns=loaded_config.pl_delay_ns,
            period_ns=loaded_config.period_ns,
            duty_percent_min=loaded_config.duty_percent_min,
            duty_percent_max=loaded_config.duty_percent_max,
            cds1_fall_us=loaded_config.cds1_fall_us,
            cds2_fall_us=loaded_config.cds2_fall_us + cds2_fall_offset_ns / 1000.0,
        )
        delay, estimates = estimate_delay(config, measured)
        residuals = {edge: estimates[edge] - value for edge, value in measured.items()}
        rms = math.sqrt(sum(value**2 for value in residuals.values()) / len(residuals))
        max_abs = max(abs(value) for value in residuals.values())
        rows.append(
            {
                "workbook": config.workbook.name,
                "index": config.index,
                "net": net,
                "estimated_delay_ns": delay,
                "cds2_fall_offset_ns": cds2_fall_offset_ns,
                "rms_error_ns": round(rms, 2),
                "max_abs_error_ns": round(max_abs, 2),
                **{
                    f"predicted_{edge}": round(estimates[edge], 2)
                    for edge in measured
                },
                **{
                    f"residual_{edge}": round(residuals[edge], 2)
                    for edge in measured
                },
            }
        )
    rows.sort(key=lambda row: (str(row["workbook"]), int(row["index"])))
    return rows


def write_result_csv(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    write_measurement_csv()
    configs = load_net_configs()
    measured_by_net: dict[str, dict[str, int]] = {}
    for net, edge, value in MEASUREMENTS:
        measured_by_net.setdefault(net, {})[edge] = value

    current_rows = build_result_rows(measured_by_net, configs, cds2_fall_offset_ns=0.0)
    rows = build_result_rows(
        measured_by_net,
        configs,
        cds2_fall_offset_ns=INFERRED_CDS2_FALL_OFFSET_NS,
    )
    write_result_csv(CURRENT_MODEL_RESULT_CSV, current_rows)
    write_result_csv(RESULT_CSV, rows)

    update_workbooks(rows)

    print(f"Wrote {MEASUREMENT_CSV}")
    print(f"Wrote {CURRENT_MODEL_RESULT_CSV}")
    print(f"Wrote {RESULT_CSV}")
    print("Updated workbooks; .xlsx.bak files were created when missing.")
    print(f"Recommended fit uses cds2_fall_offset_ns={INFERRED_CDS2_FALL_OFFSET_NS:g}.")
    for row in rows:
        print(
            f"{row['workbook']} [{row['index']}] {row['net']}: "
            f"{row['estimated_delay_ns']} ns, "
            f"RMS {row['rms_error_ns']} ns, max {row['max_abs_error_ns']} ns"
        )


if __name__ == "__main__":
    main()
