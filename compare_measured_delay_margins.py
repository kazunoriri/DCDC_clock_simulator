from __future__ import annotations

import csv
import math
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

import openpyxl


SOURCE_CLOCK_MHZ = 175.0
CONFIG_DIR = Path("設定ファイル")
BACKUP_DIR = Path("バックアップ")
CONFIG_SHEET = "config"
POWER_NET_COUNT = 8
RESULT_CSV = Path("measured_delay_margin_comparison.csv")

MEASURED_DELAYS_NS: dict[str, float] = {
    "3.3V_DIG": 44.0,
    "R2.5V_AMP": 30.0,
    "0.85V_PS": 24.0,
    "0.85V_PL": 274.0,
    "1.1V_DDRIO": 24.0,
    "3.3V_WLAN": 24.0,
    "4.3V_CHG": 42.0,
    "1.8V_DIG": 24.0,
    "5.5V_CHG": 170.0,
    "D5V_1": 130.0,
    "D5V_2": 138.0,
    "D5V_3": 142.0,
    "A6V": 186.0,
    "A3V": 62.0,
    "M15V": 330.0,
    "A20V": 306.0,
}

MARGIN_MEASUREMENTS: tuple[tuple[str, str, str, str, int], ...] = (
    ("シリアル", "PL_DCDC_CLK1", "3.3V_DIG", "CDS1↓ to SW↑", -196),
    ("シリアル", "PL_DCDC_CLK1", "3.3V_DIG", "CDS1↓ to SW↓", 103),
    ("シリアル", "PL_DCDC_CLK1", "3.3V_DIG", "CDS2↓ to SW↑", 84),
    ("シリアル", "PL_DCDC_CLK1", "3.3V_DIG", "CDS2↓ to SW↓", -97),
    ("シリアル", "PL_DCDC_CLK1", "R2.5V_AMP", "CDS1↓ to SW↑", -209),
    ("シリアル", "PL_DCDC_CLK1", "R2.5V_AMP", "CDS1↓ to SW↓", 151),
    ("シリアル", "PL_DCDC_CLK1", "R2.5V_AMP", "CDS2↓ to SW↑", 72),
    ("シリアル", "PL_DCDC_CLK1", "R2.5V_AMP", "CDS2↓ to SW↓", -42),
    ("シリアル", "PL_DCDC_CLK1", "0.85V_PS", "CDS1↓ to SW↑", -215),
    ("シリアル", "PL_DCDC_CLK1", "0.85V_PS", "CDS1↓ to SW↓", -137),
    ("シリアル", "PL_DCDC_CLK1", "0.85V_PS", "CDS2↓ to SW↑", 63),
    ("シリアル", "PL_DCDC_CLK1", "0.85V_PS", "CDS2↓ to SW↓", 142),
    ("シリアル", "PL_DCDC_CLK1", "0.85V_PL", "CDS1↓ to SW↑", 35),
    ("シリアル", "PL_DCDC_CLK1", "0.85V_PL", "CDS1↓ to SW↓", 112),
    ("シリアル", "PL_DCDC_CLK1", "0.85V_PL", "CDS2↓ to SW↑", -165),
    ("シリアル", "PL_DCDC_CLK1", "0.85V_PL", "CDS2↓ to SW↓", -88),
    ("シリアル", "PL_DCDC_CLK1", "1.1V_DDRIO", "CDS1↓ to SW↑", -215),
    ("シリアル", "PL_DCDC_CLK1", "1.1V_DDRIO", "CDS1↓ to SW↓", -114),
    ("シリアル", "PL_DCDC_CLK1", "1.1V_DDRIO", "CDS2↓ to SW↑", 65),
    ("シリアル", "PL_DCDC_CLK1", "1.1V_DDRIO", "CDS2↓ to SW↓", 166),
    ("シリアル", "PL_DCDC_CLK2", "3.3V_WLAN", "CDS1↓ to SW↑", 89),
    ("シリアル", "PL_DCDC_CLK2", "3.3V_WLAN", "CDS1↓ to SW↓", -91),
    ("シリアル", "PL_DCDC_CLK2", "3.3V_WLAN", "CDS2↓ to SW↑", -111),
    ("シリアル", "PL_DCDC_CLK2", "3.3V_WLAN", "CDS2↓ to SW↓", 189),
    ("シリアル", "PL_DCDC_CLK3", "4.3V_CHG", "CDS1↓ to SW↑", 111),
    ("シリアル", "PL_DCDC_CLK3", "4.3V_CHG", "CDS1↓ to SW↓", -194),
    ("シリアル", "PL_DCDC_CLK3", "4.3V_CHG", "CDS2↓ to SW↑", -89),
    ("シリアル", "PL_DCDC_CLK3", "4.3V_CHG", "CDS2↓ to SW↓", 87),
    ("シリアル", "PL_DCDC_CLK4", "1.8V_DIG", "CDS1↓ to SW↑", 114),
    ("シリアル", "PL_DCDC_CLK4", "1.8V_DIG", "CDS1↓ to SW↓", -202),
    ("シリアル", "PL_DCDC_CLK4", "1.8V_DIG", "CDS2↓ to SW↑", -87),
    ("シリアル", "PL_DCDC_CLK4", "1.8V_DIG", "CDS2↓ to SW↓", 77),
    ("シリアル", "PL_DCDC_CLK5", "5.5V_CHG", "CDS1↓ to SW↑", -110),
    ("シリアル", "PL_DCDC_CLK5", "5.5V_CHG", "CDS1↓ to SW↓", 238),
    ("シリアル", "PL_DCDC_CLK5", "5.5V_CHG", "CDS2↓ to SW↑", 171),
    ("シリアル", "PL_DCDC_CLK5", "5.5V_CHG", "CDS2↓ to SW↓", 39),
    ("シリアル", "PL_DCDC_CLK5", "D5V_1", "CDS1↓ to SW↑", -150),
    ("シリアル", "PL_DCDC_CLK5", "D5V_1", "CDS1↓ to SW↓", 237),
    ("シリアル", "PL_DCDC_CLK5", "D5V_1", "CDS2↓ to SW↑", 130),
    ("シリアル", "PL_DCDC_CLK5", "D5V_1", "CDS2↓ to SW↓", 37),
    ("シリアル", "PL_DCDC_CLK5", "D5V_2", "CDS1↓ to SW↑", -145),
    ("シリアル", "PL_DCDC_CLK5", "D5V_2", "CDS1↓ to SW↓", 239),
    ("シリアル", "PL_DCDC_CLK5", "D5V_2", "CDS2↓ to SW↑", 135),
    ("シリアル", "PL_DCDC_CLK5", "D5V_2", "CDS2↓ to SW↓", 39),
    ("シリアル", "PL_DCDC_CLK5", "D5V_3", "CDS1↓ to SW↑", -139),
    ("シリアル", "PL_DCDC_CLK5", "D5V_3", "CDS1↓ to SW↓", 236),
    ("シリアル", "PL_DCDC_CLK5", "D5V_3", "CDS2↓ to SW↑", 142),
    ("シリアル", "PL_DCDC_CLK5", "D5V_3", "CDS2↓ to SW↓", 36),
    ("シリアル", "PL_DCDC_CLK5", "A6V", "CDS1↓ to SW↑", -96),
    ("シリアル", "PL_DCDC_CLK5", "A6V", "CDS1↓ to SW↓", -237),
    ("シリアル", "PL_DCDC_CLK5", "A6V", "CDS2↓ to SW↑", 182),
    ("シリアル", "PL_DCDC_CLK5", "A6V", "CDS2↓ to SW↓", 41),
    ("シリアル", "PL_DCDC_CLK5", "A3V", "CDS1↓ to SW↑", -226),
    ("シリアル", "PL_DCDC_CLK5", "A3V", "CDS1↓ to SW↓", 145),
    ("シリアル", "PL_DCDC_CLK5", "A3V", "CDS2↓ to SW↑", 52),
    ("シリアル", "PL_DCDC_CLK5", "A3V", "CDS2↓ to SW↓", -57),
    ("シリアル", "PL_DCDC_CLK5", "M15V", "CDS1↓ to SW↑", 163),
    ("シリアル", "PL_DCDC_CLK5", "M15V", "CDS1↓ to SW↓", -177),
    ("シリアル", "PL_DCDC_CLK5", "M15V", "CDS2↓ to SW↑", -39),
    ("シリアル", "PL_DCDC_CLK5", "M15V", "CDS2↓ to SW↓", 103),
    ("シリアル", "PL_DCDC_CLK5", "A20V", "CDS1↓ to SW↑", 24),
    ("シリアル", "PL_DCDC_CLK5", "A20V", "CDS1↓ to SW↓", 125),
    ("シリアル", "PL_DCDC_CLK5", "A20V", "CDS2↓ to SW↑", -179),
    ("シリアル", "PL_DCDC_CLK5", "A20V", "CDS2↓ to SW↓", -84),
    ("静止画", "PL_DCDC_CLK1", "3.3V_DIG", "CDS1↓ to SW↑", -189),
    ("静止画", "PL_DCDC_CLK1", "3.3V_DIG", "CDS1↓ to SW↓", 109),
    ("静止画", "PL_DCDC_CLK1", "3.3V_DIG", "CDS2↓ to SW↑", -109),
    ("静止画", "PL_DCDC_CLK1", "3.3V_DIG", "CDS2↓ to SW↓", 190),
    ("静止画", "PL_DCDC_CLK1", "R2.5V_AMP", "CDS1↓ to SW↑", -202),
    ("静止画", "PL_DCDC_CLK1", "R2.5V_AMP", "CDS1↓ to SW↓", 160),
    ("静止画", "PL_DCDC_CLK1", "R2.5V_AMP", "CDS2↓ to SW↑", -122),
    ("静止画", "PL_DCDC_CLK1", "R2.5V_AMP", "CDS2↓ to SW↓", -236),
    ("静止画", "PL_DCDC_CLK1", "0.85V_PS", "CDS1↓ to SW↑", -211),
    ("静止画", "PL_DCDC_CLK1", "0.85V_PS", "CDS1↓ to SW↓", -82),
    ("静止画", "PL_DCDC_CLK1", "0.85V_PS", "CDS2↓ to SW↑", -130),
    ("静止画", "PL_DCDC_CLK1", "0.85V_PS", "CDS2↓ to SW↓", -57),
    ("静止画", "PL_DCDC_CLK1", "0.85V_PL", "CDS1↓ to SW↑", 42),
    ("静止画", "PL_DCDC_CLK1", "0.85V_PL", "CDS1↓ to SW↓", 119),
    ("静止画", "PL_DCDC_CLK1", "0.85V_PL", "CDS2↓ to SW↑", 122),
    ("静止画", "PL_DCDC_CLK1", "0.85V_PL", "CDS2↓ to SW↓", 198),
    ("静止画", "PL_DCDC_CLK1", "1.1V_DDRIO", "CDS1↓ to SW↑", -211),
    ("静止画", "PL_DCDC_CLK1", "1.1V_DDRIO", "CDS1↓ to SW↓", -113),
    ("静止画", "PL_DCDC_CLK1", "1.1V_DDRIO", "CDS2↓ to SW↑", -131),
    ("静止画", "PL_DCDC_CLK1", "1.1V_DDRIO", "CDS2↓ to SW↓", -32),
    ("静止画", "PL_DCDC_CLK2", "3.3V_WLAN", "CDS1↓ to SW↑", -192),
    ("静止画", "PL_DCDC_CLK2", "3.3V_WLAN", "CDS1↓ to SW↓", 107),
    ("静止画", "PL_DCDC_CLK2", "3.3V_WLAN", "CDS2↓ to SW↑", -112),
    ("静止画", "PL_DCDC_CLK2", "3.3V_WLAN", "CDS2↓ to SW↓", 186),
    ("静止画", "PL_DCDC_CLK3", "4.3V_CHG", "CDS1↓ to SW↑", -128),
    ("静止画", "PL_DCDC_CLK3", "4.3V_CHG", "CDS1↓ to SW↓", 47),
    ("静止画", "PL_DCDC_CLK3", "4.3V_CHG", "CDS2↓ to SW↑", -51),
    ("静止画", "PL_DCDC_CLK3", "4.3V_CHG", "CDS2↓ to SW↓", 124),
    ("静止画", "PL_DCDC_CLK4", "1.8V_DIG", "CDS1↓ to SW↑", -212),
    ("静止画", "PL_DCDC_CLK4", "1.8V_DIG", "CDS1↓ to SW↓", -47),
    ("静止画", "PL_DCDC_CLK4", "1.8V_DIG", "CDS2↓ to SW↑", -131),
    ("静止画", "PL_DCDC_CLK4", "1.8V_DIG", "CDS2↓ to SW↓", 29),
    ("静止画", "PL_DCDC_CLK5", "5.5V_CHG", "CDS1↓ to SW↑", 90),
    ("静止画", "PL_DCDC_CLK5", "5.5V_CHG", "CDS1↓ to SW↓", -42),
    ("静止画", "PL_DCDC_CLK5", "5.5V_CHG", "CDS2↓ to SW↑", 170),
    ("静止画", "PL_DCDC_CLK5", "5.5V_CHG", "CDS2↓ to SW↓", 38),
    ("静止画", "PL_DCDC_CLK5", "D5V_1", "CDS1↓ to SW↑", 50),
    ("静止画", "PL_DCDC_CLK5", "D5V_1", "CDS1↓ to SW↓", -43),
    ("静止画", "PL_DCDC_CLK5", "D5V_1", "CDS2↓ to SW↑", 131),
    ("静止画", "PL_DCDC_CLK5", "D5V_1", "CDS2↓ to SW↓", 37),
    ("静止画", "PL_DCDC_CLK5", "D5V_2", "CDS1↓ to SW↑", 55),
    ("静止画", "PL_DCDC_CLK5", "D5V_2", "CDS1↓ to SW↓", -41),
    ("静止画", "PL_DCDC_CLK5", "D5V_2", "CDS2↓ to SW↑", 135),
    ("静止画", "PL_DCDC_CLK5", "D5V_2", "CDS2↓ to SW↓", 39),
    ("静止画", "PL_DCDC_CLK5", "D5V_3", "CDS1↓ to SW↑", 61),
    ("静止画", "PL_DCDC_CLK5", "D5V_3", "CDS1↓ to SW↓", -44),
    ("静止画", "PL_DCDC_CLK5", "D5V_3", "CDS2↓ to SW↑", 142),
    ("静止画", "PL_DCDC_CLK5", "D5V_3", "CDS2↓ to SW↓", 36),
    ("静止画", "PL_DCDC_CLK5", "A6V", "CDS1↓ to SW↑", 103),
    ("静止画", "PL_DCDC_CLK5", "A6V", "CDS1↓ to SW↓", -37),
    ("静止画", "PL_DCDC_CLK5", "A6V", "CDS2↓ to SW↑", 183),
    ("静止画", "PL_DCDC_CLK5", "A6V", "CDS2↓ to SW↓", 43),
    ("静止画", "PL_DCDC_CLK5", "A3V", "CDS1↓ to SW↑", -26),
    ("静止画", "PL_DCDC_CLK5", "A3V", "CDS1↓ to SW↓", -136),
    ("静止画", "PL_DCDC_CLK5", "A3V", "CDS2↓ to SW↑", 53),
    ("静止画", "PL_DCDC_CLK5", "A3V", "CDS2↓ to SW↓", -55),
    ("静止画", "PL_DCDC_CLK5", "M15V", "CDS1↓ to SW↑", -118),
    ("静止画", "PL_DCDC_CLK5", "M15V", "CDS1↓ to SW↓", 21),
    ("静止画", "PL_DCDC_CLK5", "M15V", "CDS2↓ to SW↑", -40),
    ("静止画", "PL_DCDC_CLK5", "M15V", "CDS2↓ to SW↓", 100),
    ("静止画", "PL_DCDC_CLK5", "A20V", "CDS1↓ to SW↑", 221),
    ("静止画", "PL_DCDC_CLK5", "A20V", "CDS1↓ to SW↓", -161),
    ("静止画", "PL_DCDC_CLK5", "A20V", "CDS2↓ to SW↑", -180),
    ("静止画", "PL_DCDC_CLK5", "A20V", "CDS2↓ to SW↓", -83),
)


@dataclass(frozen=True)
class NetConfig:
    mode: str
    workbook: Path
    index: int
    clock: str
    name: str
    pl_delay_ns: float
    period_ns: float
    duty_percent: float
    delay_ns: float
    cds1_fall_us: float
    cds2_fall_us: float


def evaluate_formula(value: object, cells: dict[str, object]) -> object:
    if not isinstance(value, str) or not value.startswith("="):
        return value

    formula = value.replace(" ", "").upper()
    delay_match = re.fullmatch(
        r"=ROUND\(\(1/175\)\*([A-Z]+[1-9][0-9]*)\*10\^3,1\)",
        formula,
    )
    if delay_match:
        return round((1 / SOURCE_CLOCK_MHZ) * float(cells[delay_match.group(1)]) * 1000, 1)

    frequency_match = re.fullmatch(r"=ROUND\(175/([A-Z]+[1-9][0-9]*),2\)", formula)
    if frequency_match:
        return round(SOURCE_CLOCK_MHZ / float(cells[frequency_match.group(1)]), 2)

    return value


def workbook_paths() -> list[Path]:
    return sorted(CONFIG_DIR.glob("シリアル_48us/*.xlsx")) + sorted(
        CONFIG_DIR.glob("静止画_60us/*.xlsx")
    )


def load_workbook_data(path: Path) -> tuple[dict[str, object], dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[CONFIG_SHEET]
    rows = list(ws.iter_rows())
    headers = [
        str(cell.value).strip().lower() if cell.value is not None else ""
        for cell in rows[0]
    ]
    parameter_col = headers.index("parameter")
    value_col = headers.index("value")
    raw: dict[str, object] = {}
    cells: dict[str, object] = {}
    value_cells: dict[str, str] = {}

    for row in rows[1:]:
        key_cell = row[parameter_col]
        value_cell = row[value_col]
        if key_cell.value is None or not str(key_cell.value).strip():
            continue
        key = str(key_cell.value).strip()
        raw[key] = value_cell.value
        cells[value_cell.coordinate] = value_cell.value
        value_cells[key] = value_cell.coordinate

    wb.close()
    return {key: evaluate_formula(value, cells) for key, value in raw.items()}, value_cells


def update_workbooks_to_measured_delays() -> None:
    BACKUP_DIR.mkdir(exist_ok=True)
    for path in workbook_paths():
        data, value_cells = load_workbook_data(path)
        updates: list[tuple[str, float]] = []
        for index in range(POWER_NET_COUNT):
            name = str(data.get(f"power_net_name_{index}") or "").strip()
            if name in MEASURED_DELAYS_NS:
                updates.append((f"power_net_delay_ns_{index}", MEASURED_DELAYS_NS[name]))

        if not updates:
            continue

        backup = BACKUP_DIR / path.relative_to(CONFIG_DIR)
        backup.parent.mkdir(parents=True, exist_ok=True)
        if not backup.exists():
            shutil.copy2(path, backup)

        wb = openpyxl.load_workbook(path)
        ws = wb[CONFIG_SHEET]
        for key, value in updates:
            ws[value_cells[key]] = value
        wb.save(path)
        wb.close()


def load_configs() -> dict[tuple[str, str], NetConfig]:
    configs: dict[tuple[str, str], NetConfig] = {}
    for path in workbook_paths():
        mode = "シリアル" if "シリアル" in path.parts[-2] else "静止画"
        data, _ = load_workbook_data(path)
        period_ns = float(data["clock_div_ratio"]) / SOURCE_CLOCK_MHZ * 1000.0
        clock = str(data["pl_dcdc_clk_name"]).strip()
        for index in range(POWER_NET_COUNT):
            name = str(data.get(f"power_net_name_{index}") or "").strip()
            if not name:
                continue
            configs[(mode, name)] = NetConfig(
                mode=mode,
                workbook=path,
                index=index,
                clock=clock,
                name=name,
                pl_delay_ns=float(data["pl_dcdc_clk_delay_ns"]),
                period_ns=period_ns,
                duty_percent=float(data[f"power_net_duty_percent_{index}"]),
                delay_ns=float(data[f"power_net_delay_ns_{index}"]),
                cds1_fall_us=float(data["cds1_fall_us"]),
                cds2_fall_us=float(data["cds2_fall_us"]),
            )
    return configs


def nearest_edge_delta_ns(target_us: float, first_edge_ns: float, period_ns: float) -> float:
    target_ns = target_us * 1000.0
    nearest_index = round((target_ns - first_edge_ns) / period_ns)
    candidates = [
        first_edge_ns + index * period_ns
        for index in (nearest_index - 1, nearest_index, nearest_index + 1)
        if index >= 0
    ]
    if not candidates:
        candidates.append(first_edge_ns)
    nearest = min(candidates, key=lambda edge_ns: abs(edge_ns - target_ns))
    return nearest - target_ns


def predicted_margins(config: NetConfig) -> dict[str, float]:
    rise_ns = config.pl_delay_ns + config.delay_ns
    fall_ns = rise_ns + config.period_ns * config.duty_percent / 100.0
    return {
        "CDS1↓ to SW↑": nearest_edge_delta_ns(config.cds1_fall_us, rise_ns, config.period_ns),
        "CDS1↓ to SW↓": nearest_edge_delta_ns(config.cds1_fall_us, fall_ns, config.period_ns),
        "CDS2↓ to SW↑": nearest_edge_delta_ns(config.cds2_fall_us, rise_ns, config.period_ns),
        "CDS2↓ to SW↓": nearest_edge_delta_ns(config.cds2_fall_us, fall_ns, config.period_ns),
    }


def wrapped_diff_ns(calculated_ns: float, measured_ns: float, period_ns: float) -> float:
    return (calculated_ns - measured_ns + period_ns / 2.0) % period_ns - period_ns / 2.0


def write_comparison() -> list[dict[str, object]]:
    configs = load_configs()
    rows: list[dict[str, object]] = []
    for mode, clock, net, edge, measured_ns in MARGIN_MEASUREMENTS:
        config = configs[(mode, net)]
        calculated_ns = predicted_margins(config)[edge]
        rows.append(
            {
                "mode": mode,
                "clock": clock,
                "workbook": str(config.workbook),
                "power_net": net,
                "timing": edge,
                "delay_ns": config.delay_ns,
                "measured_margin_ns": measured_ns,
                "simulated_margin_ns": round(calculated_ns, 2),
                "diff_sim_minus_measured_ns": round(calculated_ns - measured_ns, 2),
                "wrapped_diff_sim_minus_measured_ns": round(
                    wrapped_diff_ns(calculated_ns, measured_ns, config.period_ns),
                    2,
                ),
            }
        )

    with RESULT_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    return rows


def print_summary(rows: list[dict[str, object]]) -> None:
    print(f"Wrote {RESULT_CSV}")
    for mode in ("シリアル", "静止画"):
        values = [
            float(row["wrapped_diff_sim_minus_measured_ns"])
            for row in rows
            if row["mode"] == mode
        ]
        rms = math.sqrt(sum(value * value for value in values) / len(values))
        max_abs = max(abs(value) for value in values)
        print(f"{mode}: RMS={rms:.2f} ns, max_abs={max_abs:.2f} ns")

    worst = sorted(
        rows,
        key=lambda row: abs(float(row["wrapped_diff_sim_minus_measured_ns"])),
        reverse=True,
    )[:12]
    print("Worst wrapped diffs:")
    for row in worst:
        print(
            f"{row['mode']} {row['clock']} {row['power_net']} {row['timing']}: "
            f"sim={row['simulated_margin_ns']} ns, measured={row['measured_margin_ns']} ns, "
            f"diff={row['wrapped_diff_sim_minus_measured_ns']} ns"
        )


def main() -> None:
    rows = write_comparison()
    print_summary(rows)


if __name__ == "__main__":
    main()
