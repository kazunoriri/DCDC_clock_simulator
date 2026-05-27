from __future__ import annotations

import math
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pyqtgraph as pg
from PyQt6 import QtCore, QtGui, QtWidgets


US_PER_NS = 0.001
POWER_MARGIN_RANGE_NS = 250.0
DUTY_MARGIN_SAMPLE_STEP_PERCENT = 0.1
TITLE_FONT_SIZE_PT = 10
APP_FONT_SIZE_PT = 9
OUTER_MARGIN_PX = 12
GRAPH_SPACING_PX = 24
ROW_SPACING_PX = 14
CONTROL_ROW_HEIGHT_PX = 72
COMPACT_WINDOW_HEIGHT_PX = 600
MAX_PL_DELAY_NS = 250.0
PLOT_BOTTOM_AXIS_HEIGHT_PX = 62
TIMING_LEFT_AXIS_WIDTH_PX = 120
MARGIN_LEFT_AXIS_WIDTH_PX = 75
SWEEP_LEFT_AXIS_WIDTH_PX = 75
MAX_POWER_TICK_LABEL_CHARS = 12
CONFIG_PATH = Path(__file__).with_name("timing_config_debug.xlsx")
CONFIG_SHEET_NAME = "config"
CONFIG_PARAMETER_COLUMN = "parameter"
CONFIG_VALUE_COLUMN = "value"
POWER_NET_COUNT = 8
TIMING_FIXED_ROW_COUNT = 3 + POWER_NET_COUNT
EXCEL_CELL_PATTERN = re.compile(r"^[A-Z]+[1-9][0-9]*$")
CONFIG_REQUIRED_KEYS = [
    "pl_dcdc_clk_name",
    "pl_dcdc_clk_delay_ns",
    "gate_period_us",
    "cds1_rise_us",
    "cds1_fall_us",
    "cds2_rise_us",
    "cds2_fall_us",
    "clock_div_ratio",
    *[f"power_net_name_{index}" for index in range(POWER_NET_COUNT)],
    *[f"power_net_delay_ns_{index}" for index in range(POWER_NET_COUNT)],
    *[f"power_net_duty_percent_{index}_min" for index in range(POWER_NET_COUNT)],
    *[f"power_net_duty_percent_{index}_max" for index in range(POWER_NET_COUNT)],
]


@dataclass(frozen=True)
class PlotTheme:
    background: str
    text: str
    title: str
    axis: str
    border: str
    grid_alpha: float
    waveform: str
    marker: str
    zero_line: str
    cds1: str
    cds2: str
    control_background: str
    control_border: str
    control_text: str
    selection_background: str
    button_background: str
    button_hover: str


SCREEN_THEME = PlotTheme(
    background="#000000",
    text="#b8c3d1",
    title="#b8c3d1",
    axis="#808080",
    border="#606060",
    grid_alpha=0.34,
    waveform="#ffff00",
    marker="#8a8a8a",
    zero_line="#d0d7de",
    cds1="#58a6ff",
    cds2="#f0883e",
    control_background="#555555",
    control_border="#606060",
    control_text="#f1f5f9",
    selection_background="#1f6feb",
    button_background="#555555",
    button_hover="#666666",
)


COPY_THEME = PlotTheme(
    background="#ffffff",
    text="#20242a",
    title="#20242a",
    axis="#4b5563",
    border="#8b949e",
    grid_alpha=0.22,
    waveform="#0057d8",
    marker="#7a7f87",
    zero_line="#30363d",
    cds1="#0057d8",
    cds2="#f97316",
    control_background="#ffffff",
    control_border="#8b949e",
    control_text="#20242a",
    selection_background="#9ec5ff",
    button_background="#e5e7eb",
    button_hover="#d1d5db",
)


@dataclass(frozen=True)
class PulseSignal:
    name: str
    pulses_us: tuple[tuple[float, float], ...]


@dataclass(frozen=True)
class PowerNet:
    name: str
    delay_ns: float
    duty_percent_min: float
    duty_percent_max: float

    @property
    def duty(self) -> float:
        return self.duty_percent_min / 100.0

    @property
    def has_duty_range(self) -> bool:
        return not math.isclose(self.duty_percent_min, self.duty_percent_max)


@dataclass(frozen=True)
class TimingConfig:
    pl_dcdc_clk_name: str = "PL_DCDC_CLK1"
    pl_dcdc_clk_delay_ns: float = 0.0
    power_net_name_0: str = "3.3V_DIG"
    power_net_name_1: str = "3.3V_DIG_2"
    power_net_name_2: str = "3.3V_DIG_3"
    power_net_name_3: str = "3.3V_DIG_4"
    power_net_name_4: str = "3.3V_DIG_5"
    power_net_name_5: str = "3.3V_DIG_6"
    power_net_name_6: str = ""
    power_net_name_7: str = ""
    power_net_delay_ns_0: float | None = 0.0
    power_net_delay_ns_1: float | None = 10.0
    power_net_delay_ns_2: float | None = 20.0
    power_net_delay_ns_3: float | None = 30.0
    power_net_delay_ns_4: float | None = 40.0
    power_net_delay_ns_5: float | None = 50.0
    power_net_delay_ns_6: float | None = None
    power_net_delay_ns_7: float | None = None
    power_net_duty_percent_0_min: float | None = 35.0
    power_net_duty_percent_1_min: float | None = 42.5
    power_net_duty_percent_2_min: float | None = 50.0
    power_net_duty_percent_3_min: float | None = 57.5
    power_net_duty_percent_4_min: float | None = 65.0
    power_net_duty_percent_5_min: float | None = 72.5
    power_net_duty_percent_6_min: float | None = None
    power_net_duty_percent_7_min: float | None = None
    power_net_duty_percent_0_max: float | None = 35.0
    power_net_duty_percent_1_max: float | None = 42.5
    power_net_duty_percent_2_max: float | None = 50.0
    power_net_duty_percent_3_max: float | None = 57.5
    power_net_duty_percent_4_max: float | None = 65.0
    power_net_duty_percent_5_max: float | None = 72.5
    power_net_duty_percent_6_max: float | None = None
    power_net_duty_percent_7_max: float | None = None
    gate_period_us: float = 60.0
    cds1_rise_us: float = 1.0
    cds1_fall_us: float = 15.0
    cds2_rise_us: float = 30.0
    cds2_fall_us: float = 40.0
    clock_div_ratio: int = 84


def load_timing_config(path: Path | None = None) -> tuple[TimingConfig | None, str | None]:
    if path is None:
        path = CONFIG_PATH

    if not path.exists():
        return None, None

    try:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except OSError as exc:
        return None, f"Excel file could not be read: {exc}"
    except Exception as exc:
        return None, f"Excel parse error: {exc}"

    if CONFIG_SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        return None, f'Sheet "{CONFIG_SHEET_NAME}" not found.'

    worksheet = workbook[CONFIG_SHEET_NAME]
    rows = worksheet.iter_rows()
    try:
        header = next(rows)
    except StopIteration:
        workbook.close()
        return None, "Excel config sheet is empty."

    normalized_header = [
        str(cell.value).strip().lower() if cell.value is not None else ""
        for cell in header
    ]
    try:
        parameter_index = normalized_header.index(CONFIG_PARAMETER_COLUMN)
        value_index = normalized_header.index(CONFIG_VALUE_COLUMN)
    except ValueError:
        workbook.close()
        return None, 'Expected columns: "parameter", "value".'

    cell_values = {}
    raw_data = {}
    for row in rows:
        parameter_cell = row[parameter_index] if parameter_index < len(row) else None
        value_cell = row[value_index] if value_index < len(row) else None
        parameter = parameter_cell.value if parameter_cell is not None else None
        if parameter is None or not str(parameter).strip():
            continue
        value = value_cell.value if value_cell is not None else None
        coordinate = getattr(value_cell, "coordinate", None)
        if coordinate is not None:
            cell_values[coordinate] = value
        raw_data[str(parameter).strip()] = value
    workbook.close()

    def evaluate_excel_formula(value: object) -> object:
        if not isinstance(value, str) or not value.startswith("="):
            return value

        formula = value.replace(" ", "").upper()
        delay_match = re.fullmatch(
            r"=ROUND\(\(1/175\)\*([A-Z]+[1-9][0-9]*)\*10\^3,1\)",
            formula,
        )
        if delay_match:
            ref = delay_match.group(1)
            if not EXCEL_CELL_PATTERN.match(ref):
                return value
            try:
                return round((1 / 175) * float(cell_values[ref]) * 10**3, 1)
            except (KeyError, TypeError, ValueError):
                return value

        frequency_match = re.fullmatch(
            r"=ROUND\(175/([A-Z]+[1-9][0-9]*),2\)",
            formula,
        )
        if frequency_match:
            ref = frequency_match.group(1)
            if not EXCEL_CELL_PATTERN.match(ref):
                return value
            try:
                return round(175 / float(cell_values[ref]), 2)
            except (KeyError, TypeError, ValueError, ZeroDivisionError):
                return value

        return value

    data = {
        key: evaluate_excel_formula(value)
        for key, value in raw_data.items()
    }

    missing_keys = [key for key in CONFIG_REQUIRED_KEYS if key not in data]
    if missing_keys:
        return None, "Missing required parameter(s): " + ", ".join(missing_keys)

    def parse_str(key: str) -> str:
        value = data.get(key)
        return str(value) if value is not None else ""

    def parse_positive_float(key: str) -> float | None:
        try:
            value = float(data[key])
        except (TypeError, ValueError):
            return None
        return value if math.isfinite(value) and value >= 0.0 else None

    def parse_duty_percent(key: str) -> float | None:
        try:
            value = float(data[key])
        except (TypeError, ValueError):
            return None
        return value if math.isfinite(value) and 0.0 < value <= 100.0 else None

    def parse_duty_percent_range(index: int) -> tuple[float | None, float | None]:
        min_key = f"power_net_duty_percent_{index}_min"
        max_key = f"power_net_duty_percent_{index}_max"
        duty_min = parse_duty_percent(min_key)
        duty_max = parse_duty_percent(max_key)
        if (duty_min is None) != (duty_max is None):
            raise ValueError(f"{min_key} and {max_key} must be specified together.")
        if duty_min is not None and duty_max is not None and duty_min > duty_max:
            raise ValueError(f"{min_key} must be <= {max_key}.")
        return duty_min, duty_max

    def parse_required_float(key: str, *, positive: bool = False) -> tuple[float, str | None]:
        try:
            value = float(data[key])
        except (TypeError, ValueError):
            return 0.0, f"{key} must be a number."
        if not math.isfinite(value):
            return 0.0, f"{key} must be finite."
        if positive and value <= 0.0:
            return 0.0, f"{key} must be greater than 0."
        return value, None

    def parse_required_int(key: str, *, minimum: int) -> tuple[int, str | None]:
        try:
            value = int(data[key])
        except (TypeError, ValueError):
            return 0, f"{key} must be an integer."
        if value < minimum:
            return 0, f"{key} must be >= {minimum}."
        return value, None

    try:
        gate_period_us, error = parse_required_float("gate_period_us", positive=True)
        if error:
            return None, error
        cds1_rise_us, error = parse_required_float("cds1_rise_us")
        if error:
            return None, error
        cds1_fall_us, error = parse_required_float("cds1_fall_us")
        if error:
            return None, error
        cds2_rise_us, error = parse_required_float("cds2_rise_us")
        if error:
            return None, error
        cds2_fall_us, error = parse_required_float("cds2_fall_us")
        if error:
            return None, error
        clock_div_ratio, error = parse_required_int("clock_div_ratio", minimum=1)
        if error:
            return None, error
        pl_dcdc_clk_delay_ns, error = parse_required_float("pl_dcdc_clk_delay_ns")
        if error:
            return None, error
        if pl_dcdc_clk_delay_ns < 0.0 or pl_dcdc_clk_delay_ns > MAX_PL_DELAY_NS:
            return None, f"pl_dcdc_clk_delay_ns must be between 0 and {MAX_PL_DELAY_NS:g}."

        values = {
            "pl_dcdc_clk_name": parse_str("pl_dcdc_clk_name"),
            "pl_dcdc_clk_delay_ns": pl_dcdc_clk_delay_ns,
            "gate_period_us": gate_period_us,
            "cds1_rise_us": cds1_rise_us,
            "cds1_fall_us": cds1_fall_us,
            "cds2_rise_us": cds2_rise_us,
            "cds2_fall_us": cds2_fall_us,
            "clock_div_ratio": clock_div_ratio,
        }
        for index in range(POWER_NET_COUNT):
            values[f"power_net_name_{index}"] = parse_str(f"power_net_name_{index}")
            values[f"power_net_delay_ns_{index}"] = parse_positive_float(
                f"power_net_delay_ns_{index}"
            )
            duty_min, duty_max = parse_duty_percent_range(index)
            values[f"power_net_duty_percent_{index}_min"] = duty_min
            values[f"power_net_duty_percent_{index}_max"] = duty_max
    except (TypeError, ValueError) as exc:
        return None, f"Configuration error: {exc}"

    if not values["pl_dcdc_clk_name"].strip():
        return None, "pl_dcdc_clk_name must not be blank."
    if values["cds1_rise_us"] >= values["cds1_fall_us"]:
        return None, "cds1_rise_us must be less than cds1_fall_us."
    if values["cds2_rise_us"] >= values["cds2_fall_us"]:
        return None, "cds2_rise_us must be less than cds2_fall_us."
    for index in range(POWER_NET_COUNT):
        name = values[f"power_net_name_{index}"].strip()
        delay_ns = values[f"power_net_delay_ns_{index}"]
        duty_min = values[f"power_net_duty_percent_{index}_min"]
        duty_max = values[f"power_net_duty_percent_{index}_max"]
        if name and delay_ns is not None and (duty_min is None or duty_max is None):
            return (
                None,
                f"power_net_duty_percent_{index}_min/max must be numbers "
                "between 0 and 100.",
            )

    return TimingConfig(**values), None


def get_preferred_font_family() -> str:
    app = QtWidgets.QApplication.instance()
    if app is None:
        return "Sans Serif"

    families = set(QtGui.QFontDatabase.families())
    for family in (
        "Yu Gothic UI",
        "Hiragino Sans",
        "Hiragino Kaku Gothic ProN",
        "Arial Unicode MS",
        "Meiryo",
    ):
        if family in families:
            return family
    return app.font().family()


class TimeAxisItem(pg.AxisItem):
    def tickStrings(self, values, scale, spacing):  # noqa: N802
        if spacing >= 1:
            digits = 0
        elif spacing >= 0.1:
            digits = 1
        elif spacing >= 0.01:
            digits = 2
        elif spacing >= 0.001:
            digits = 3
        else:
            digits = 6

        labels = []
        for value in values:
            text = f"{value * scale:.{digits}f}"
            labels.append(text.rstrip("0").rstrip(".") if "." in text else text)
        return labels


def shorten_tick_label(text: str, max_chars: int = MAX_POWER_TICK_LABEL_CHARS) -> str:
    if len(text) <= max_chars:
        return text
    if max_chars <= 1:
        return text[:max_chars]
    return text[: max_chars - 1] + "…"


def pulse_steps(
    pulses_us: tuple[tuple[float, float], ...],
    x_min_us: float,
    x_max_us: float,
    low: float,
    high: float,
) -> tuple[list[float], list[float]]:
    points: list[tuple[float, float]] = [(x_min_us, low)]
    for start_us, end_us in sorted(pulses_us):
        points.extend(
            [
                (start_us, low),
                (start_us, high),
                (end_us, high),
                (end_us, low),
            ]
        )
    points.append((x_max_us, low))
    x, y = zip(*points)
    return list(x), list(y)


def clock_steps(
    start_us: float,
    end_us: float,
    period_us: float,
    duty: float,
    low: float,
    high: float,
) -> tuple[list[float], list[float]]:
    x = [start_us]
    y = [low]
    t = start_us
    high_width = period_us * duty

    while t < end_us:
        rise = t
        fall = min(t + high_width, end_us)
        next_rise = min(t + period_us, end_us)
        x.extend([rise, rise, fall, fall, next_rise])
        y.extend([low, high, high, low, low])
        t += period_us

    return x, y


class TimingDiagram(QtWidgets.QMainWindow):
    def __init__(
        self,
        *,
        export_mode: bool = False,
        include_copy_button: bool = True,
    ) -> None:
        super().__init__()
        self.export_mode = export_mode
        self.include_copy_button = include_copy_button
        self.theme = COPY_THEME if export_mode else SCREEN_THEME
        self.setWindowTitle("DCDC Clock Timing")
        self.resize(1760, COMPACT_WINDOW_HEIGHT_PX)
        if not self.export_mode:
            self.move(0, 0)
            self.setAcceptDrops(True)

        self.x_min_us = -2.0
        self.row_gap = 1.05
        self.amplitude = 0.74
        self.source_clock_mhz = 175.0
        self.pl_clk1_delay_ns = 0.0
        self.power_duty_display_mode = "min"
        self.config: TimingConfig | None = None
        self.has_timing_config = False
        startup_config, startup_error = load_timing_config()
        self.configure_from_config(startup_config)
        self.startup_error = startup_error
        self.font_family = get_preferred_font_family()
        self.app_font = QtGui.QFont(self.font_family, APP_FONT_SIZE_PT)
        self.baselines: dict[str, float] = {}
        self.pl_delay_status_label: QtWidgets.QLabel | None = None
        self.power_duty_mode_combo: QtWidgets.QComboBox | None = None
        self.copy_image_button: QtWidgets.QPushButton | None = None

        pg.setConfigOptions(antialias=True)

        central = QtWidgets.QWidget()
        central.setStyleSheet(
            f"background: {self.theme.background}; color: {self.theme.text};"
        )
        layout = QtWidgets.QGridLayout(central)
        layout.setContentsMargins(
            OUTER_MARGIN_PX,
            OUTER_MARGIN_PX,
            OUTER_MARGIN_PX,
            OUTER_MARGIN_PX,
        )
        layout.setHorizontalSpacing(GRAPH_SPACING_PX)
        layout.setVerticalSpacing(ROW_SPACING_PX)
        layout.setColumnStretch(0, 100)
        layout.setColumnStretch(1, 100)
        layout.setColumnStretch(2, 100)
        layout.setRowStretch(0, 1)
        layout.setRowStretch(1, 0)
        layout.setRowMinimumHeight(1, CONTROL_ROW_HEIGHT_PX)

        self.plot = pg.PlotWidget(axisItems={"bottom": TimeAxisItem(orientation="bottom")})
        self.margin_plot = pg.PlotWidget()
        self.delay_sweep_plot = pg.PlotWidget()
        layout.addWidget(self.plot, 0, 0)
        layout.addWidget(self.margin_plot, 0, 1)
        layout.addWidget(self.delay_sweep_plot, 0, 2)
        layout.addWidget(self.create_control_panel(), 1, 0)
        layout.addWidget(self.create_margin_legend_with_button(), 1, 1)
        layout.addWidget(QtWidgets.QWidget(), 1, 2)
        self.setCentralWidget(central)

        self.plot.setBackground(self.theme.background)
        self.plot.showGrid(x=True, y=True, alpha=self.theme.grid_alpha)
        self.plot.setTitle(
            "Timing waveform",
            color=self.theme.title,
            size=f"{TITLE_FONT_SIZE_PT}pt",
        )
        self.plot.setLabel(
            "bottom",
            "time [us]",
            **self.axis_label_style(),
        )
        self.plot.setMouseEnabled(x=True, y=False)
        self.plot.setMenuEnabled(False)
        self.plot.getPlotItem().getViewBox().setBorder(
            pg.mkPen(self.theme.border, width=1)
        )

        bottom_axis = self.plot.getAxis("bottom")
        left_axis = self.plot.getAxis("left")
        for axis in (bottom_axis, left_axis):
            axis.setPen(pg.mkPen(self.theme.axis))
            axis.setTextPen(pg.mkPen(self.theme.text))
            axis.setStyle(
                tickFont=self.app_font,
                tickTextOffset=8,
                autoExpandTextSpace=True,
            )
        bottom_axis.setHeight(PLOT_BOTTOM_AXIS_HEIGHT_PX)
        left_axis.setWidth(TIMING_LEFT_AXIS_WIDTH_PX)

        self.margin_plot.setBackground(self.theme.background)
        self.margin_plot.showGrid(x=False, y=True, alpha=self.theme.grid_alpha)
        self.margin_plot.setLabel(
            "left",
            "delta [ns]",
            **self.axis_label_style(),
        )
        self.margin_plot.setLabel(
            "bottom",
            "電源名",
            **self.axis_label_style(),
        )
        self.margin_plot.setTitle(
            "CDS1,2↓ to switching edge delta",
            color=self.theme.title,
            size=f"{TITLE_FONT_SIZE_PT}pt",
        )
        self.margin_plot.setMouseEnabled(x=False, y=False)
        self.margin_plot.setMenuEnabled(False)
        self.margin_plot.getPlotItem().getViewBox().setBorder(
            pg.mkPen(self.theme.border, width=1)
        )
        margin_bottom_axis = self.margin_plot.getAxis("bottom")
        margin_left_axis = self.margin_plot.getAxis("left")
        for axis in (margin_bottom_axis, margin_left_axis):
            axis.setPen(pg.mkPen(self.theme.axis))
            axis.setTextPen(pg.mkPen(self.theme.text))
            axis.setStyle(
                tickFont=self.app_font,
                tickTextOffset=6,
                autoExpandTextSpace=True,
            )
        margin_bottom_axis.setHeight(PLOT_BOTTOM_AXIS_HEIGHT_PX)
        margin_left_axis.setWidth(MARGIN_LEFT_AXIS_WIDTH_PX)

        self.delay_sweep_plot.setBackground(self.theme.background)
        self.delay_sweep_plot.showGrid(x=True, y=True, alpha=self.theme.grid_alpha)
        self.delay_sweep_plot.setLabel(
            "left",
            "delta [ns]",
            **self.axis_label_style(),
        )
        self.delay_sweep_plot.setLabel(
            "bottom",
            f"{self.pl_dcdc_clk_name} delay [ns]",
            **self.axis_label_style(),
        )
        self.delay_sweep_plot.setTitle(
            "delta vs delay",
            color=self.theme.title,
            size=f"{TITLE_FONT_SIZE_PT}pt",
        )
        self.delay_sweep_plot.setMouseEnabled(x=True, y=False)
        self.delay_sweep_plot.setMenuEnabled(False)
        self.delay_sweep_plot.getPlotItem().getViewBox().setBorder(
            pg.mkPen(self.theme.border, width=1)
        )
        sweep_bottom_axis = self.delay_sweep_plot.getAxis("bottom")
        sweep_left_axis = self.delay_sweep_plot.getAxis("left")
        for axis in (sweep_bottom_axis, sweep_left_axis):
            axis.setPen(pg.mkPen(self.theme.axis))
            axis.setTextPen(pg.mkPen(self.theme.text))
            axis.setStyle(
                tickFont=self.app_font,
                tickTextOffset=6,
                autoExpandTextSpace=True,
            )
        sweep_bottom_axis.setHeight(PLOT_BOTTOM_AXIS_HEIGHT_PX)
        sweep_left_axis.setWidth(SWEEP_LEFT_AXIS_WIDTH_PX)

        self.draw()
        if self.startup_error and not self.export_mode:
            QtCore.QTimer.singleShot(
                0,
                lambda: self.show_config_error(
                    "Config error",
                    f"{CONFIG_PATH.name}: {self.startup_error}",
                ),
            )

    def configure_from_config(self, config: TimingConfig | None) -> None:
        self.config = config
        self.has_timing_config = config is not None
        active_config = config or TimingConfig()
        self.pl_clk1_delay_ns = (
            active_config.pl_dcdc_clk_delay_ns if self.has_timing_config else 0.0
        )
        self.gate_period_us = active_config.gate_period_us
        self.x_max_us = self.gate_period_us + 2.0
        self.pl_dcdc_clk_name = active_config.pl_dcdc_clk_name
        power_net_names = [
            getattr(active_config, f"power_net_name_{index}")
            for index in range(POWER_NET_COUNT)
        ]
        power_net_delays_ns = [
            getattr(active_config, f"power_net_delay_ns_{index}")
            for index in range(POWER_NET_COUNT)
        ]
        power_net_duty_percent_mins = [
            getattr(active_config, f"power_net_duty_percent_{index}_min")
            for index in range(POWER_NET_COUNT)
        ]
        power_net_duty_percent_maxs = [
            getattr(active_config, f"power_net_duty_percent_{index}_max")
            for index in range(POWER_NET_COUNT)
        ]
        self.power_nets = [
            PowerNet(name.strip(), delay_ns, duty_percent_min, duty_percent_max)
            for name, delay_ns, duty_percent_min, duty_percent_max in zip(
                power_net_names,
                power_net_delays_ns,
                power_net_duty_percent_mins,
                power_net_duty_percent_maxs,
                strict=True,
            )
            if (
                name.strip()
                and delay_ns is not None
                and duty_percent_min is not None
                and duty_percent_max is not None
            )
        ]
        if not self.has_timing_config:
            self.power_nets = []
        self.power_net_names = [power_net.name for power_net in self.power_nets]
        self.power_net_name_0 = self.power_net_names[0] if self.power_net_names else ""
        self.clock_div_ratio = active_config.clock_div_ratio
        self.cds1_start_us = active_config.cds1_rise_us
        self.cds1_end_us = active_config.cds1_fall_us
        self.cds2_start_us = active_config.cds2_rise_us
        self.cds2_end_us = active_config.cds2_fall_us
        self.pl_delay_candidates_ns = self.create_pl_delay_candidates_ns()
        self.pl_clk1_delay_ns = self.nearest_pl_delay_candidate_ns(
            self.pl_clk1_delay_ns
        )
        self.row_names = (
            [
                "CDS1",
                "CDS2",
                self.pl_dcdc_clk_name,
                *self.power_net_names,
            ]
            if self.has_timing_config
            else []
        )

    def axis_label_style(self) -> dict[str, str]:
        return {
            "color": self.theme.text,
            "font-size": f"{APP_FONT_SIZE_PT}pt",
            "font-family": self.font_family,
        }

    def create_control_panel(self) -> QtWidgets.QWidget:
        panel = QtWidgets.QWidget()
        panel.setStyleSheet(
            f"""
            QLabel {{
                color: {self.theme.text};
                font-family: "{self.font_family}";
                font-size: {APP_FONT_SIZE_PT}pt;
            }}
            QLineEdit {{
                background: {self.theme.control_background};
                border: 1px solid {self.theme.control_border};
                color: {self.theme.control_text};
                font-family: "{self.font_family}";
                font-size: {APP_FONT_SIZE_PT}pt;
                padding: 2px 6px;
                selection-background-color: {self.theme.selection_background};
            }}
            QComboBox {{
                background: {self.theme.control_background};
                border: 1px solid {self.theme.control_border};
                color: {self.theme.control_text};
                font-family: "{self.font_family}";
                font-size: {APP_FONT_SIZE_PT}pt;
                padding: 2px 6px;
                selection-background-color: {self.theme.selection_background};
            }}
            QLabel#result {{
                color: {self.theme.text};
            }}
            """
        )
        self.control_panel = panel
        layout = QtWidgets.QHBoxLayout(panel)
        layout.setContentsMargins(8, 6, 8, 8)
        layout.setSpacing(12)

        self.pl_clk1_delay_combo = self.create_delay_combo()
        self.power_duty_mode_combo = self.create_duty_mode_combo()
        self.control_widgets_by_row: dict[str, QtWidgets.QWidget] = {}

        layout.addWidget(self.create_clock_editor(parent=panel))
        layout.addStretch(1)
        return panel

    def create_clock_editor(self, parent: QtWidgets.QWidget) -> QtWidgets.QWidget:
        editor = QtWidgets.QWidget(parent)
        layout = QtWidgets.QGridLayout(editor)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setHorizontalSpacing(10)
        layout.setVerticalSpacing(4)
        layout.setColumnStretch(0, 0)
        layout.setColumnStretch(1, 0)
        layout.setColumnStretch(2, 1)

        self.pl_delay_status_label = QtWidgets.QLabel(self.delay_status_text())
        duty_label = QtWidgets.QLabel("power duty display")

        layout.addWidget(self.pl_delay_status_label, 0, 0)
        layout.addWidget(self.pl_clk1_delay_combo, 0, 1)
        layout.addWidget(duty_label, 1, 0)
        layout.addWidget(self.power_duty_mode_combo, 1, 1)
        editor.adjustSize()
        return editor

    def create_margin_legend(self) -> QtWidgets.QWidget:
        legend = QtWidgets.QWidget()
        legend.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Ignored,
            QtWidgets.QSizePolicy.Policy.Preferred,
        )
        layout = QtWidgets.QVBoxLayout(legend)
        layout.setContentsMargins(8, 6, 8, 8)
        layout.setSpacing(2)

        entries = [
            [
                ("△", "CDS1↓ to SW↑", self.theme.cds1),
                ("▽", "CDS1↓ to SW↓", self.theme.cds1),
            ],
            [
                ("▲", "CDS2↓ to SW↑", self.theme.cds2),
                ("▼", "CDS2↓ to SW↓", self.theme.cds2),
            ],
        ]
        for row_entries in entries:
            row = QtWidgets.QWidget(legend)
            row_layout = QtWidgets.QHBoxLayout(row)
            row_layout.setContentsMargins(0, 0, 0, 0)
            row_layout.setSpacing(18)
            row_layout.addStretch(1)
            for marker, text, color in row_entries:
                label = QtWidgets.QLabel(f"{marker} {text}")
                label.setStyleSheet(
                    f'color: {color}; font-family: "{self.font_family}"; '
                    f"font-size: {APP_FONT_SIZE_PT}pt;"
                )
                row_layout.addWidget(label)
            row_layout.addStretch(1)
            layout.addWidget(row)
        return legend

    def create_margin_legend_with_button(self) -> QtWidgets.QWidget:
        """凡例とcopy imageボタンを縦に並べたパネル（中グラフの下に配置）"""
        panel = QtWidgets.QWidget()
        panel.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Ignored,
            QtWidgets.QSizePolicy.Policy.Preferred,
        )
        layout = QtWidgets.QVBoxLayout(panel)
        layout.setContentsMargins(8, 6, 8, 8)
        layout.setSpacing(4)

        entries = [
            [
                ("△", "CDS1↓ to SW↑", self.theme.cds1),
                ("▽", "CDS1↓ to SW↓", self.theme.cds1),
            ],
            [
                ("▲", "CDS2↓ to SW↑", self.theme.cds2),
                ("▼", "CDS2↓ to SW↓", self.theme.cds2),
            ],
        ]
        for row_entries in entries:
            row = QtWidgets.QWidget(panel)
            row_layout = QtWidgets.QHBoxLayout(row)
            row_layout.setContentsMargins(0, 0, 0, 0)
            row_layout.setSpacing(18)
            row_layout.addStretch(1)
            for marker, text, color in row_entries:
                label = QtWidgets.QLabel(f"{marker} {text}")
                label.setStyleSheet(
                    f'color: {color}; font-family: "{self.font_family}"; '
                    f"font-size: {APP_FONT_SIZE_PT}pt;"
                )
                row_layout.addWidget(label)
            row_layout.addStretch(1)
            layout.addWidget(row)

        if self.include_copy_button:
            button = QtWidgets.QPushButton("copy image")
            self.copy_image_button = button
            button.setFixedWidth(104)
            button.setFixedHeight(30)
            button.setStyleSheet(
                f"""
                QPushButton {{
                    background: {self.theme.button_background};
                    border: 1px solid {self.theme.control_border};
                    color: {self.theme.control_text};
                    font-family: "{self.font_family}";
                    font-size: {APP_FONT_SIZE_PT}pt;
                    padding: 4px 12px;
                }}
                QPushButton:hover {{
                    background: {self.theme.button_hover};
                }}
                QPushButton:pressed {{
                    background: {self.theme.control_border};
                    padding-top: 5px;
                    padding-bottom: 3px;
                }}
                """
            )
            button.clicked.connect(self.copy_image_to_clipboard)
            button_row = QtWidgets.QWidget(panel)
            button_layout = QtWidgets.QHBoxLayout(button_row)
            button_layout.setContentsMargins(0, 0, 0, 0)
            button_layout.addStretch(1)
            button_layout.addWidget(button)
            layout.addWidget(button_row)

        return panel

    def create_copy_button_panel(self) -> QtWidgets.QWidget:
        panel = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(panel)
        layout.setContentsMargins(8, 6, 8, 8)
        layout.addStretch(1)

        button = QtWidgets.QPushButton("copy image")
        self.copy_image_button = button
        button.setFixedWidth(104)
        button.setFixedHeight(30)
        button.setStyleSheet(
            f"""
            QPushButton {{
                background: {self.theme.button_background};
                border: 1px solid {self.theme.control_border};
                color: {self.theme.control_text};
                font-family: "{self.font_family}";
                font-size: {APP_FONT_SIZE_PT}pt;
                padding: 4px 12px;
            }}
            QPushButton:hover {{
                background: {self.theme.button_hover};
            }}
            QPushButton:pressed {{
                background: {self.theme.control_border};
                padding-top: 5px;
                padding-bottom: 3px;
            }}
            """
        )
        button.clicked.connect(self.copy_image_to_clipboard)
        button_row = QtWidgets.QWidget(panel)
        button_layout = QtWidgets.QHBoxLayout(button_row)
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.addStretch(1)
        button_layout.addWidget(button)
        layout.addWidget(button_row)
        return panel

    def create_pl_delay_candidates_ns(self) -> list[float]:
        candidates: list[float] = []
        source_period_ns = 1000.0 / self.source_clock_mhz
        step = 0
        while True:
            delay_ns = step * source_period_ns
            if delay_ns > MAX_PL_DELAY_NS + 1e-9:
                break
            candidates.append(delay_ns)
            step += 1
        return candidates

    def create_delay_combo(self) -> QtWidgets.QComboBox:
        combo = QtWidgets.QComboBox()
        combo.setFixedWidth(86)
        for delay_ns in self.pl_delay_candidates_ns:
            combo.addItem(f"{delay_ns:.1f}", delay_ns)
        self.select_nearest_delay_combo_index(combo, self.pl_clk1_delay_ns)
        combo.setEnabled(self.has_timing_config)
        combo.currentIndexChanged.connect(self.apply_inputs)
        return combo

    def create_duty_mode_combo(self) -> QtWidgets.QComboBox:
        combo = QtWidgets.QComboBox()
        combo.setFixedWidth(86)
        combo.addItem("min", "min")
        combo.addItem("max", "max")
        combo.setEnabled(self.has_timing_config)
        combo.currentIndexChanged.connect(self.apply_inputs)
        return combo

    def nearest_pl_delay_candidate_ns(self, delay_ns: float) -> float:
        if not self.pl_delay_candidates_ns:
            return delay_ns
        return min(
            self.pl_delay_candidates_ns,
            key=lambda candidate_ns: abs(candidate_ns - delay_ns),
        )

    def select_nearest_delay_combo_index(
        self,
        combo: QtWidgets.QComboBox,
        delay_ns: float,
    ) -> None:
        if combo.count() == 0:
            return
        best_index = min(
            range(combo.count()),
            key=lambda index: abs(float(combo.itemData(index)) - delay_ns),
        )
        combo.setCurrentIndex(best_index)

    @staticmethod
    def triangle_symbol(up: bool) -> QtGui.QPainterPath:
        path = QtGui.QPainterPath()
        if up:
            points = [
                QtCore.QPointF(0.0, -0.55),
                QtCore.QPointF(0.55, 0.55),
                QtCore.QPointF(-0.55, 0.55),
            ]
        else:
            points = [
                QtCore.QPointF(0.0, 0.55),
                QtCore.QPointF(0.55, -0.55),
                QtCore.QPointF(-0.55, -0.55),
            ]
        path.moveTo(points[0])
        path.lineTo(points[1])
        path.lineTo(points[2])
        path.closeSubpath()
        return path

    def draw(self) -> None:
        previous_x_range: list[float] | None = None
        previous_y_range: list[float] | None = None
        if self.baselines:
            previous_x_range, previous_y_range = self.plot.viewRange()

        self.plot.clear()
        if not self.has_timing_config:
            self.draw_empty_plots()
            return

        signals = [
            PulseSignal("CDS1", ((self.cds1_start_us, self.cds1_end_us),)),
            PulseSignal("CDS2", ((self.cds2_start_us, self.cds2_end_us),)),
        ]

        waveform_pen = pg.mkPen(self.theme.waveform, width=1)
        marker_pen = pg.mkPen(
            self.theme.marker,
            width=1,
            style=QtCore.Qt.PenStyle.DashLine,
        )

        baselines = {
            name: (TIMING_FIXED_ROW_COUNT - index - 1) * self.row_gap
            for index, name in enumerate(self.row_names)
        }
        self.baselines = baselines
        half_amp = self.amplitude * 0.5
        clock_period_us = self.clock_div_ratio / self.source_clock_mhz

        self.plot.getAxis("left").setTicks(
            [[(baselines[name], name) for name in self.row_names]]
        )

        for signal in signals:
            base = baselines[signal.name]
            x, y = pulse_steps(
                signal.pulses_us,
                0.0,
                self.gate_period_us,
                base - half_amp,
                base + half_amp,
            )
            self.plot.plot(x, y, pen=waveform_pen)

        pl_base = baselines[self.pl_dcdc_clk_name]
        pl_start_us = self.pl_clk1_delay_ns * US_PER_NS
        pl_x, pl_y = clock_steps(
            start_us=pl_start_us,
            end_us=self.gate_period_us,
            period_us=clock_period_us,
            duty=0.5,
            low=pl_base - half_amp,
            high=pl_base + half_amp,
        )
        if pl_start_us > 0.0:
            pl_x = [0.0, pl_start_us] + pl_x
            pl_y = [pl_base - half_amp, pl_base - half_amp] + pl_y
        self.plot.plot(pl_x, pl_y, pen=waveform_pen)

        for power_net in self.power_nets:
            dig_base = baselines[power_net.name]
            dig_start_us = (self.pl_clk1_delay_ns + power_net.delay_ns) * US_PER_NS
            dig_x, dig_y = clock_steps(
                start_us=dig_start_us,
                end_us=self.gate_period_us,
                period_us=clock_period_us,
                duty=self.display_power_net_duty(power_net),
                low=dig_base - half_amp,
                high=dig_base + half_amp,
            )
            if dig_start_us > 0.0:
                dig_x = [0.0, dig_start_us] + dig_x
                dig_y = [dig_base - half_amp, dig_base - half_amp] + dig_y
            self.plot.plot(dig_x, dig_y, pen=waveform_pen)
        self.draw_margin_plot(clock_period_us)
        self.draw_delay_sweep_plot(clock_period_us)

        self.add_time_markers(
            [0, self.cds1_end_us, self.cds2_end_us, self.gate_period_us],
            marker_pen,
        )

        if previous_x_range is None or previous_y_range is None:
            self.plot.setXRange(self.x_min_us, self.x_max_us, padding=0)
            self.plot.setYRange(
                -0.75,
                (TIMING_FIXED_ROW_COUNT - 1) * self.row_gap + 0.75,
                padding=0,
            )
        else:
            self.plot.setXRange(previous_x_range[0], previous_x_range[1], padding=0)
            self.plot.setYRange(previous_y_range[0], previous_y_range[1], padding=0)

    def draw_empty_plots(self) -> None:
        self.baselines = {}
        self.plot.getAxis("left").setTicks([[]])
        self.plot.setXRange(0.0, 1.0, padding=0)
        self.plot.setYRange(0.0, 1.0, padding=0)

        self.margin_plot.clear()
        self.margin_plot.getAxis("bottom").setTicks([[]])
        self.margin_plot.setXRange(0.0, 1.0, padding=0)
        self.margin_plot.setYRange(0.0, 1.0, padding=0)

        self.delay_sweep_plot.clear()
        self.delay_sweep_plot.setXRange(0.0, 1.0, padding=0)
        self.delay_sweep_plot.setYRange(0.0, 1.0, padding=0)

    def draw_margin_plot(self, clock_period_us: float) -> None:
        self.margin_plot.clear()
        self.update_margin_axis_ticks()
        power_net_count = len(self.power_nets)
        self.margin_plot.setXRange(
            -0.5,
            max(power_net_count - 0.5, 0.5),
            padding=0,
        )
        self.margin_plot.setYRange(
            -POWER_MARGIN_RANGE_NS,
            POWER_MARGIN_RANGE_NS,
            padding=0,
        )

        self.margin_plot.addItem(
            pg.InfiniteLine(pos=0, angle=0, pen=pg.mkPen(self.theme.zero_line, width=2))
        )
        for value_ns in (-POWER_MARGIN_RANGE_NS, POWER_MARGIN_RANGE_NS):
            self.margin_plot.addItem(
                pg.InfiniteLine(
                    pos=value_ns,
                    angle=0,
                    pen=pg.mkPen(
                        self.theme.marker,
                        width=1,
                        style=QtCore.Qt.PenStyle.DashLine,
                    ),
                )
            )

        deltas = self.calculate_power_edge_deltas_ns(clock_period_us)
        self.update_margin_tooltip(deltas)
        series = [
            ("cds1_rise", -0.18, self.triangle_symbol(up=True), self.theme.cds1, None),
            (
                "cds1_fall",
                -0.06,
                self.triangle_symbol(up=False),
                self.theme.cds1,
                None,
            ),
            (
                "cds2_rise",
                0.06,
                self.triangle_symbol(up=True),
                self.theme.cds2,
                self.theme.cds2,
            ),
            (
                "cds2_fall",
                0.18,
                self.triangle_symbol(up=False),
                self.theme.cds2,
                self.theme.cds2,
            ),
        ]
        for key, x_offset, symbol, color, brush_color in series:
            if key.endswith("_fall"):
                self.add_margin_range_bars(deltas, key, x_offset, color)
                x_values: list[float] = []
                y_values: list[float] = []
                for index, entry in enumerate(deltas):
                    x_values.extend([index + x_offset, index + x_offset])
                    y_values.extend(
                        [
                            entry[f"{key}_duty_min"],
                            entry[f"{key}_duty_max"],
                        ]
                    )
            else:
                x_values = [index + x_offset for index in range(len(deltas))]
                y_values = [entry[key] for entry in deltas]
            self.margin_plot.plot(
                x_values,
                y_values,
                pen=None,
                symbol=symbol,
                symbolSize=9,
                symbolPen=pg.mkPen(color, width=1.5),
                symbolBrush=pg.mkBrush(brush_color) if brush_color else None,
            )

    def update_margin_axis_ticks(self) -> None:
        self.margin_plot.getAxis("bottom").setTicks(
            [
                [
                    (index, power_net.name)
                    for index, power_net in enumerate(self.power_nets)
                ]
            ]
        )
        self.margin_plot.getAxis("bottom").setStyle(
            tickFont=self.margin_tick_label_font(),
            tickTextOffset=6,
            autoExpandTextSpace=True,
        )

    def display_power_net_duty(self, power_net: PowerNet) -> float:
        if self.power_duty_display_mode == "max":
            return power_net.duty_percent_max / 100.0
        return power_net.duty_percent_min / 100.0

    def margin_tick_label_max_chars(self, power_net_count: int) -> int:
        if power_net_count <= 0:
            return MAX_POWER_TICK_LABEL_CHARS

        plot_width = max(self.margin_plot.width(), 1)
        usable_width = max(plot_width - MARGIN_LEFT_AXIS_WIDTH_PX - 24, 1)
        slot_width = usable_width / power_net_count
        char_width = max(self.app_font_metrics().horizontalAdvance("M"), 1)
        dynamic_chars = max(4, int(slot_width / char_width) - 1)
        return min(MAX_POWER_TICK_LABEL_CHARS, dynamic_chars)

    def margin_tick_label_font(self) -> QtGui.QFont:
        font = QtGui.QFont(self.font_family, APP_FONT_SIZE_PT)
        power_net_count = len(self.power_nets)
        if power_net_count <= 0:
            return font

        plot_width = max(self.margin_plot.width(), 1)
        usable_width = max(plot_width - MARGIN_LEFT_AXIS_WIDTH_PX - 24, 1)
        slot_width = usable_width / power_net_count
        metrics = QtGui.QFontMetrics(font)
        longest_width = max(
            metrics.horizontalAdvance(power_net.name)
            for power_net in self.power_nets
        )
        if longest_width <= slot_width:
            return font

        scaled_size = max(6, int(APP_FONT_SIZE_PT * slot_width / longest_width))
        font.setPointSize(scaled_size)
        return font

    def app_font_metrics(self) -> QtGui.QFontMetrics:
        return QtGui.QFontMetrics(self.app_font)

    def add_margin_range_bars(
        self,
        deltas: list[dict[str, float]],
        key: str,
        x_offset: float,
        color: str,
    ) -> None:
        pen = pg.mkPen(color, width=1.6)
        for index, entry in enumerate(deltas):
            x = index + x_offset
            for y_min, y_max in entry[f"{key}_segments"]:
                self.margin_plot.plot(
                    [x, x],
                    [y_min, y_max],
                    pen=pen,
                )

    def update_margin_tooltip(self, deltas: list[dict[str, float]]) -> None:
        if not deltas:
            self.margin_plot.setToolTip("")
            return

        lines = []
        for power_net, entry in zip(self.power_nets, deltas, strict=True):
            lines.append(power_net.name)
            for label, key in (
                ("CDS1↓ to SW↓", "cds1_fall"),
                ("CDS2↓ to SW↓", "cds2_fall"),
            ):
                lines.append(
                    f"  {label}: {entry[f'{key}_min']:.1f}〜"
                    f"{entry[f'{key}_max']:.1f} ns, "
                    f"worst {entry[f'{key}_worst_abs']:.1f} ns"
                )
        self.margin_plot.setToolTip("\n".join(lines))

    def draw_delay_sweep_plot(self, clock_period_us: float) -> None:
        self.delay_sweep_plot.clear()
        self.delay_sweep_plot.setXRange(-1.0, MAX_PL_DELAY_NS + 1.0, padding=0)
        self.delay_sweep_plot.setYRange(
            -POWER_MARGIN_RANGE_NS,
            POWER_MARGIN_RANGE_NS,
            padding=0,
        )
        self.delay_sweep_plot.addItem(
            pg.InfiniteLine(pos=0, angle=0, pen=pg.mkPen(self.theme.zero_line, width=2))
        )
        for value_ns in (-POWER_MARGIN_RANGE_NS, POWER_MARGIN_RANGE_NS):
            self.delay_sweep_plot.addItem(
                pg.InfiniteLine(
                    pos=value_ns,
                    angle=0,
                    pen=pg.mkPen(
                        self.theme.marker,
                        width=1,
                        style=QtCore.Qt.PenStyle.DashLine,
                    ),
                )
            )

        series = [
            ("cds1_rise", -0.18, self.triangle_symbol(up=True), self.theme.cds1, None),
            (
                "cds1_fall",
                -0.06,
                self.triangle_symbol(up=False),
                self.theme.cds1,
                None,
            ),
            (
                "cds2_rise",
                0.06,
                self.triangle_symbol(up=True),
                self.theme.cds2,
                self.theme.cds2,
            ),
            (
                "cds2_fall",
                0.18,
                self.triangle_symbol(up=False),
                self.theme.cds2,
                self.theme.cds2,
            ),
        ]
        power_offsets = [
            (index - (len(self.power_nets) - 1) * 0.5) * 0.16
            for index in range(len(self.power_nets))
        ]

        series_points: list[
            tuple[list[float], list[float], QtGui.QPainterPath, str, str | None]
        ] = []

        for key, series_offset, symbol, color, brush_color in series:
            if not key.endswith("_fall"):
                continue
            range_pen = pg.mkPen(QtGui.QColor(color).lighter(115), width=0.8)
            for delay_ns in self.pl_delay_candidates_ns:
                deltas = self.calculate_power_edge_deltas_ns(
                    clock_period_us,
                    pl_delay_ns=delay_ns,
                )
                for power_index, entry in enumerate(deltas):
                    x = delay_ns + series_offset + power_offsets[power_index]
                    for y_min, y_max in entry[f"{key}_segments"]:
                        self.delay_sweep_plot.plot(
                            [x, x],
                            [y_min, y_max],
                            pen=range_pen,
                        )

        for key, series_offset, symbol, color, brush_color in series:
            x_values: list[float] = []
            y_values: list[float] = []
            for delay_ns in self.pl_delay_candidates_ns:
                deltas = self.calculate_power_edge_deltas_ns(
                    clock_period_us,
                    pl_delay_ns=delay_ns,
                )
                for power_index, entry in enumerate(deltas):
                    x = delay_ns + series_offset + power_offsets[power_index]
                    if key.endswith("_fall"):
                        x_values.extend([x, x])
                        y_values.extend(
                            [
                                entry[f"{key}_duty_min"],
                                entry[f"{key}_duty_max"],
                            ]
                        )
                    else:
                        x_values.append(x)
                        y_values.append(entry[key])
            series_points.append((x_values, y_values, symbol, color, brush_color))

        for x_values, y_values, symbol, color, brush_color in series_points:
            self.delay_sweep_plot.plot(
                x_values,
                y_values,
                pen=None,
                symbol=symbol,
                symbolSize=8,
                symbolPen=pg.mkPen(color, width=1.5),
                symbolBrush=pg.mkBrush(brush_color) if brush_color else None,
            )

    def calculate_power_edge_deltas_ns(
        self,
        clock_period_us: float,
        pl_delay_ns: float | None = None,
    ) -> list[dict[str, float]]:
        if pl_delay_ns is None:
            pl_delay_ns = self.pl_clk1_delay_ns

        deltas: list[dict[str, float]] = []
        for power_net in self.power_nets:
            rise_start_us = (pl_delay_ns + power_net.delay_ns) * US_PER_NS
            cds1_fall = self.sample_fall_margin_range_ns(
                self.cds1_end_us,
                rise_start_us,
                clock_period_us,
                power_net,
            )
            cds2_fall = self.sample_fall_margin_range_ns(
                self.cds2_end_us,
                rise_start_us,
                clock_period_us,
                power_net,
            )
            deltas.append(
                {
                    "cds1_rise": self.nearest_clock_edge_delta_ns(
                        self.cds1_end_us,
                        rise_start_us,
                        clock_period_us,
                    ),
                    "cds1_fall": cds1_fall["duty_min"],
                    "cds1_fall_duty_min": cds1_fall["duty_min"],
                    "cds1_fall_duty_max": cds1_fall["duty_max"],
                    "cds1_fall_min": cds1_fall["min"],
                    "cds1_fall_max": cds1_fall["max"],
                    "cds1_fall_worst_abs": cds1_fall["worst_abs"],
                    "cds1_fall_segments": cds1_fall["segments"],
                    "cds2_rise": self.nearest_clock_edge_delta_ns(
                        self.cds2_end_us,
                        rise_start_us,
                        clock_period_us,
                    ),
                    "cds2_fall": cds2_fall["duty_min"],
                    "cds2_fall_duty_min": cds2_fall["duty_min"],
                    "cds2_fall_duty_max": cds2_fall["duty_max"],
                    "cds2_fall_min": cds2_fall["min"],
                    "cds2_fall_max": cds2_fall["max"],
                    "cds2_fall_worst_abs": cds2_fall["worst_abs"],
                    "cds2_fall_segments": cds2_fall["segments"],
                }
            )
        return deltas

    def sample_fall_margin_range_ns(
        self,
        target_us: float,
        rise_start_us: float,
        clock_period_us: float,
        power_net: PowerNet,
    ) -> dict[str, float]:
        samples = self.duty_percent_samples(
            power_net.duty_percent_min,
            power_net.duty_percent_max,
        )
        margins = [
            self.nearest_clock_edge_delta_ns(
                target_us,
                rise_start_us + clock_period_us * duty_percent / 100.0,
                clock_period_us,
            )
            for duty_percent in samples
        ]
        duty_min_margin = margins[0]
        duty_max_margin = margins[-1]
        segments = self.split_margin_segments_ns(
            margins,
            period_ns=clock_period_us / US_PER_NS,
        )
        return {
            "duty_min": duty_min_margin,
            "duty_max": duty_max_margin,
            "min": min(margins),
            "max": max(margins),
            "worst_abs": min(abs(margin) for margin in margins),
            "segments": segments,
        }

    def split_margin_segments_ns(
        self,
        margins: list[float],
        *,
        period_ns: float,
    ) -> list[tuple[float, float]]:
        if not margins:
            return []

        segments: list[list[float]] = [[margins[0]]]
        jump_threshold_ns = period_ns * 0.5
        for previous, current in zip(margins, margins[1:]):
            if abs(current - previous) > jump_threshold_ns:
                segments.append([current])
            else:
                segments[-1].append(current)

        return [(min(segment), max(segment)) for segment in segments]

    def duty_percent_samples(self, duty_min: float, duty_max: float) -> list[float]:
        if math.isclose(duty_min, duty_max):
            return [duty_min]

        samples = [duty_min]
        value = duty_min + DUTY_MARGIN_SAMPLE_STEP_PERCENT
        while value < duty_max:
            samples.append(value)
            value += DUTY_MARGIN_SAMPLE_STEP_PERCENT
        samples.append(duty_max)
        return samples

    def align_control_panel(self) -> None:
        if not self.baselines:
            return

        view_box = self.plot.getPlotItem().getViewBox()
        for row_name, widget in self.control_widgets_by_row.items():
            scene_pos = view_box.mapViewToScene(
                QtCore.QPointF(self.x_min_us, self.baselines[row_name])
            )
            plot_pos = self.plot.mapFromScene(scene_pos)
            widget.adjustSize()
            y = int(plot_pos.y() - widget.height() / 2)
            widget.move(0, y)

    def add_time_markers(self, values_us: list[float], marker_pen: QtGui.QPen) -> None:
        for value_us in values_us:
            line = pg.InfiniteLine(
                pos=value_us,
                angle=90,
                pen=marker_pen,
            )
            self.plot.addItem(line)

    def nearest_clock_edge_delta_ns(
        self,
        target_us: float,
        first_edge_us: float,
        period_us: float,
    ) -> float:
        nearest_index = round((target_us - first_edge_us) / period_us)
        candidates = []
        for index in (nearest_index - 1, nearest_index, nearest_index + 1):
            if index >= 0:
                candidates.append(first_edge_us + index * period_us)

        if not candidates:
            candidates.append(first_edge_us)

        nearest_edge_us = min(candidates, key=lambda edge_us: abs(edge_us - target_us))
        return (nearest_edge_us - target_us) / US_PER_NS

    def apply_inputs(self) -> None:
        if not self.has_timing_config:
            return
        pl_clk1_delay_ns = self.parse_delay_edit()
        if pl_clk1_delay_ns is None:
            return

        self.pl_clk1_delay_ns = pl_clk1_delay_ns
        self.power_duty_display_mode = self.parse_duty_mode()
        self.draw()

    def refresh_config_widgets(self) -> None:
        if self.pl_delay_status_label is not None:
            self.pl_delay_status_label.setText(self.delay_status_text())
        if hasattr(self, "pl_clk1_delay_combo"):
            self.pl_clk1_delay_combo.blockSignals(True)
            self.pl_clk1_delay_combo.clear()
            for delay_ns in self.pl_delay_candidates_ns:
                self.pl_clk1_delay_combo.addItem(f"{delay_ns:.1f}", delay_ns)
            self.select_nearest_delay_combo_index(
                self.pl_clk1_delay_combo,
                self.pl_clk1_delay_ns,
            )
            self.pl_clk1_delay_combo.blockSignals(False)
            self.pl_clk1_delay_combo.setEnabled(self.has_timing_config)
        if self.power_duty_mode_combo is not None:
            self.power_duty_mode_combo.blockSignals(True)
            self.select_duty_mode_combo_index()
            self.power_duty_mode_combo.blockSignals(False)
            self.power_duty_mode_combo.setEnabled(self.has_timing_config)
        if hasattr(self, "delay_sweep_plot"):
            self.delay_sweep_plot.setLabel(
                "bottom",
                f"{self.pl_dcdc_clk_name} delay [ns]"
                if self.has_timing_config
                else "",
                **self.axis_label_style(),
            )

    def apply_config_from_path(self, path: Path) -> bool:
        config, error = load_timing_config(path)
        if error:
            self.show_config_error("Config error", f"{path.name}: {error}")
            return False
        if config is None:
            self.show_config_error("Config error", f"{path.name}: file not found.")
            return False

        self.configure_from_config(config)
        self.refresh_config_widgets()
        self.draw()
        return True

    def show_config_error(self, title: str, message: str) -> None:
        if self.export_mode:
            return
        QtWidgets.QMessageBox.warning(self, title, message)

    def first_excel_drop_path(self, event: QtGui.QDropEvent) -> Path | None:
        for url in event.mimeData().urls():
            if not url.isLocalFile():
                continue
            path = Path(url.toLocalFile())
            if path.suffix.lower() == ".xlsx":
                return path
        return None

    def dragEnterEvent(self, event: QtGui.QDragEnterEvent) -> None:  # noqa: N802
        if self.first_excel_drop_path(event) is not None:
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event: QtGui.QDropEvent) -> None:  # noqa: N802
        path = self.first_excel_drop_path(event)
        if path is None:
            self.show_config_error("Config error", "Drop an Excel file.")
            event.ignore()
            return
        if self.apply_config_from_path(path):
            event.acceptProposedAction()
        else:
            event.ignore()

    def parse_delay_edit(self) -> float | None:
        delay_ns = self.pl_clk1_delay_combo.currentData()
        if delay_ns is None:
            return None
        return float(delay_ns)

    def parse_duty_mode(self) -> str:
        if self.power_duty_mode_combo is None:
            return self.power_duty_display_mode
        mode = self.power_duty_mode_combo.currentData()
        return str(mode) if mode in {"min", "max"} else "min"

    def select_duty_mode_combo_index(self) -> None:
        if self.power_duty_mode_combo is None:
            return
        for index in range(self.power_duty_mode_combo.count()):
            if self.power_duty_mode_combo.itemData(index) == self.power_duty_display_mode:
                self.power_duty_mode_combo.setCurrentIndex(index)
                return

    def delay_status_text(self) -> str:
        if not self.has_timing_config:
            return "Drop Excel config"
        return f"{self.pl_dcdc_clk_name}: delay [ns]"

    def plot_view_ranges(self) -> dict[str, tuple[list[float], list[float]]]:
        return {
            "plot": self.plot.viewRange(),
            "margin_plot": self.margin_plot.viewRange(),
            "delay_sweep_plot": self.delay_sweep_plot.viewRange(),
        }

    def apply_plot_view_ranges(
        self,
        ranges: dict[str, tuple[list[float], list[float]]],
    ) -> None:
        for attr_name, (x_range, y_range) in ranges.items():
            plot_widget = getattr(self, attr_name)
            plot_widget.setXRange(x_range[0], x_range[1], padding=0)
            plot_widget.setYRange(y_range[0], y_range[1], padding=0)

    def copy_image_to_clipboard(self) -> None:
        # ボタンを即座に "copied" に変更・無効化（タイマーはコピー完了後にセット）
        button = getattr(self, "copy_image_button", None)
        if button is not None:
            button.setText("copied")
            button.setEnabled(False)
            QtWidgets.QApplication.processEvents()

        copy_window: TimingDiagram | None = None
        try:
            copy_window = TimingDiagram(
                export_mode=True,
                include_copy_button=False,
            )
            copy_window.setAttribute(
                QtCore.Qt.WidgetAttribute.WA_DontShowOnScreen,
                True,
            )
            copy_window.resize(self.size())
            copy_window.configure_from_config(self.config)
            copy_window.refresh_config_widgets()
            copy_window.pl_clk1_delay_combo.setCurrentIndex(
                self.pl_clk1_delay_combo.currentIndex()
            )
            if copy_window.power_duty_mode_combo is not None:
                copy_window.power_duty_mode_combo.setCurrentIndex(
                    self.power_duty_mode_combo.currentIndex()
                    if self.power_duty_mode_combo is not None
                    else 0
                )
            copy_window.power_duty_display_mode = self.power_duty_display_mode
            copy_window.pl_clk1_delay_ns = self.pl_clk1_delay_ns
            copy_window.draw()
            copy_window.apply_plot_view_ranges(self.plot_view_ranges())
            copy_window.show()
            QtWidgets.QApplication.processEvents()

            # 中グラフ（margin_plot）の ViewBox 範囲をキャプチャ
            # （グラフタイトル・縦軸ラベル・横軸タイトルを除外、電源名ラベルは含む）
            central_widget = copy_window.centralWidget()
            margin_plot_widget = copy_window.margin_plot

            # ViewBox の矩形を centralWidget のウィジェット座標系に変換
            view_box = copy_window.margin_plot.getPlotItem().getViewBox()
            vb_scene_rect = view_box.mapRectToScene(view_box.boundingRect())
            vb_tl_scene = vb_scene_rect.topLeft()
            vb_br_scene = vb_scene_rect.bottomRight()
            vb_tl_local = margin_plot_widget.mapFromScene(vb_tl_scene)
            vb_br_local = margin_plot_widget.mapFromScene(vb_br_scene)
            vb_tl = central_widget.mapFromGlobal(
                margin_plot_widget.mapToGlobal(
                    QtCore.QPoint(int(vb_tl_local.x()), int(vb_tl_local.y()))
                )
            )
            vb_br = central_widget.mapFromGlobal(
                margin_plot_widget.mapToGlobal(
                    QtCore.QPoint(int(vb_br_local.x()), int(vb_br_local.y()))
                )
            )

            # 縦軸ラベル（"delta [ns]"、90度回転）の幅だけ除外し、数値目盛りは含める
            left_axis = copy_window.margin_plot.getAxis("left")
            left_label_width = int(left_axis.label.boundingRect().height())
            margin_pos = margin_plot_widget.mapTo(central_widget, QtCore.QPoint(0, 0))
            crop_left = margin_pos.x() + left_label_width
            crop_top = vb_tl.y()
            # 右枠線が切れないよう +2px（centralWidget幅を超えないようクランプ）
            crop_right = min(vb_br.x() + 2, central_widget.width())

            # 電源名ラベルを含めるため ViewBox 下端から目盛りラベル分だけ下に伸ばす
            # PLOT_BOTTOM_AXIS_HEIGHT_PX = 目盛りラベル + 軸タイトルの合計高さ
            bottom_axis = copy_window.margin_plot.getAxis("bottom")
            label_height = int(bottom_axis.label.boundingRect().height())
            tick_label_height = PLOT_BOTTOM_AXIS_HEIGHT_PX - label_height
            crop_bottom = min(vb_br.y() + tick_label_height, central_widget.height())

            crop_rect = QtCore.QRect(
                crop_left,
                crop_top,
                crop_right - crop_left,
                crop_bottom - crop_top,
            )
            pixmap = central_widget.grab(crop_rect)
            QtWidgets.QApplication.clipboard().setPixmap(pixmap)
        finally:
            if copy_window is not None:
                copy_window.close()
                copy_window.deleteLater()

        # コピー処理完了後に700msタイマーをセット
        if button is not None:
            QtCore.QTimer.singleShot(
                700,
                lambda: (button.setText("copy image"), button.setEnabled(True)),
            )

    def show_copy_feedback(self) -> None:
        button = getattr(self, "copy_image_button", None)
        if button is None:
            return
        button.setText("copied")
        button.setEnabled(False)

        def restore() -> None:
            button.setText("copy image")
            button.setEnabled(True)

        QtCore.QTimer.singleShot(700, restore)

    def resizeEvent(self, event: QtGui.QResizeEvent) -> None:  # noqa: N802
        super().resizeEvent(event)
        QtCore.QTimer.singleShot(0, self.align_control_panel)
        QtCore.QTimer.singleShot(0, self.update_margin_axis_ticks)


def main() -> None:
    app = QtWidgets.QApplication(sys.argv)
    window = TimingDiagram()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
